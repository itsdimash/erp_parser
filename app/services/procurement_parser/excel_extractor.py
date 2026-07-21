

from __future__ import annotations

import re
from dataclasses import dataclass
from typing import Optional

import openpyxl

from .config import ParserConfig
from .models import Product
from .text_utils import normalize_ws, normalize_name, map_headers, parse_quantity

# Same class of stop words as product_extractor.py's PDF path.
_STOP_WORDS = (
    "итого", "всего", "total", "subtotal", "sum", "сумма прописью",
    "раздел", "section",
)

# A real article code looks like "522-101-1212008" / "245-401-0201".
_CODE_RE = re.compile(r"\d{2,3}-\d{2,3}-\d+")


@dataclass
class ExcelColumnMap:
    """1-based column indices."""
    item_no: int = 1
    code: int = 2
    name: int = 3
    unit: int = 7
    quantity: int = 8
    price: int = 9
    total: int = 10


class ExcelProductExtractor:
    def __init__(
        self,
        column_map: Optional[ExcelColumnMap] = None,
        sheet_name: Optional[str] = None,
        parser_config: Optional[ParserConfig] = None,
    ):
        self.column_map = column_map
        self.sheet_name = sheet_name
        self.parser_config = parser_config or ParserConfig()

    def extract(self, xlsx_path: str) -> list[Product]:
        wb = openpyxl.load_workbook(xlsx_path, data_only=True)
        ws = wb[self.sheet_name] if self.sheet_name else wb[wb.sheetnames[0]]

        rows = [
            [normalize_ws(str(c.value)) if c.value is not None else "" for c in row]
            for row in ws.iter_rows(min_row=1, max_row=ws.max_row)
        ]
        rows = [row for row in rows if any(cell for cell in row)]
        if not rows:
            return []

        if self.column_map is not None:
            # Use explicit 1-based Excel mapping when provided.
            header_map = {
                "item_no": self.column_map.item_no - 1,
                "code": self.column_map.code - 1,
                "name": self.column_map.name - 1,
                "unit": self.column_map.unit - 1,
                "quantity": self.column_map.quantity - 1,
                "price": self.column_map.price - 1,
                "total": self.column_map.total - 1,
            }
            header_row_index = 0
        else:
            header_row_index, header_map = self._find_best_header(rows)
            if header_map is None:
                # fallback to defaults for the first row if no header row is found.
                header_map = {
                    "item_no": 0,
                    "code": 1,
                    "name": 2,
                    "unit": 6,
                    "quantity": 7,
                    "price": 8,
                    "total": 9,
                }
                header_row_index = 0

        cm = ExcelColumnMap(
            item_no=header_map.get("item_no", 0) + 1,
            code=header_map.get("code", 1) + 1,
            name=header_map.get("name", 2) + 1,
            unit=header_map.get("unit", 6) + 1,
            quantity=header_map.get("quantity", 7) + 1,
            price=header_map.get("price", 8) + 1,
            total=header_map.get("total", 9) + 1,
        )

        products: list[Product] = []
        for row in rows[header_row_index + 1 :]:
            code_cell = row[cm.code - 1] if cm.code - 1 < len(row) else ""
            name_cell = row[cm.name - 1] if cm.name - 1 < len(row) else ""
            qty_cell = row[cm.quantity - 1] if cm.quantity - 1 < len(row) else ""

            if not name_cell and not code_cell:
                continue
            if name_cell.isdigit() and code_cell.isdigit():
                continue

            joined = (name_cell + " " + code_cell).lower()
            if any(sw in joined for sw in _STOP_WORDS):
                continue

            name = normalize_name(name_cell)
            if not name:
                continue

            quantity = parse_quantity(qty_cell)
            products.append(Product(name=name, quantity=quantity))

        return self._dedupe(products)

    def _find_best_header(self, rows: list[list[str]]) -> tuple[int, dict[str, int] | None]:
        best_map: dict[str, int] | None = None
        best_score = -1
        best_idx = 0
        for idx, row in enumerate(rows[:8]):
            if not any(row):
                continue
            cmap = map_headers(row, self.parser_config)
            if not cmap:
                continue
            score = len(cmap) + (1 if "name" in cmap or "description" in cmap else 0)
            if score > best_score:
                best_score = score
                best_map = cmap
                best_idx = idx
        return best_idx, best_map

    @staticmethod
    def _to_number(value) -> Optional[float]:
        if value is None:
            return None
        if isinstance(value, (int, float)):
            return float(value)
        s = normalize_ws(str(value)).replace(" ", "").replace(",", ".")
        try:
            return float(s)
        except ValueError:
            return None

    @staticmethod
    def _dedupe(products: list[Product]) -> list[Product]:
        seen: set[tuple] = set()
        out: list[Product] = []
        for p in products:
            key = (p.name.lower(), p.quantity)
            if key in seen:
                continue
            seen.add(key)
            out.append(p)
        return out
