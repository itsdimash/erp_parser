"""Generate a minimal two-column Excel with Product and Quantity."""
from __future__ import annotations
from typing import Iterable

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from .models import Product

class ExcelGenerator:
    def generate(self, products: Iterable[Product], output_path: str) -> str:
        wb = Workbook()
        ws = wb.active
        ws.title = "Products"

        # Header style: bold white text, blue fill, centered
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")

        for col, title in enumerate(("Product", "Quantity"), start=1):
            cell = ws.cell(row=1, column=col, value=title)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment

        # Make the header row a bit taller so the centered text has room
        ws.row_dimensions[1].height = 20

        # This is the strict setting that stops text from bleeding out of the box
        # It forces the text to wrap inside the cell.
        wrap_alignment = Alignment(wrap_text=True, vertical="top")

        for r, p in enumerate(products, start=2):
            c1 = ws.cell(row=r, column=1, value=p.name)
            c2 = ws.cell(row=r, column=2, value=(p.quantity if p.quantity is not None else ""))
            
            # Apply the wrap text alignment to every cell
            c1.alignment = wrap_alignment
            c2.alignment = wrap_alignment

        # Increase the width of the Product column to be extremely wide
        ws.column_dimensions['A'].width = 120 
        ws.column_dimensions['B'].width = 20

        wb.save(output_path)
        return output_path